import openpyxl as xl

wb = xl.load_workbook('sheet.xlsx')
sheet = wb['Sheet1']
cell = sheet.cell(1,1)
print(sheet.max_row)
ctMarks=list(range(4))
for row in range(sheet.min_row+1,sheet.max_row+1):
    for i in range(3,7):
        marks = sheet.cell(row,i)
        ctMarks[i - 3] = marks.value

wb.save('sheet2.xlsx')
